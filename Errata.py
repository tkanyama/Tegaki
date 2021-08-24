# -*- coding: utf-8 -*-
'''

'''

import wx
import wx.grid
import openpyxl

FRAME_WIDTH = 420
FRAME_HEIGHT = 400
CTR_HEIGHT = 40
LAST_HEIGHT = 30
BUTTON_WIDTH = 60
BUTTON_HEIGHT = 40

UNDO = 1001
INS = 1002
DUP = 1003
CUT = 1004
COPY = 1005
PASTE = 1006


class Errata(wx.Frame):
    def __init__(self, parent, id):
        wx.Frame.__init__(self, parent, id, '正誤表の編集', size=(FRAME_WIDTH, FRAME_HEIGHT),
                          style=wx.STATIC_BORDER | wx.SYSTEM_MENU | wx.CAPTION | wx.CLIP_CHILDREN | wx.STAY_ON_TOP)
        self.Center()
        (self.frame_w, self.frame_h) = self.GetSize()  # ウィンドウのサイズの読み込み

        # =======================[ メイン　パネル ]=========================================
        self.mainPanel = wx.Panel(self, wx.ID_ANY, pos=(0, 0), size=(self.frame_w, self.frame_h))

        self.ctrl_panel = wx.Panel(self.mainPanel,wx.ID_ANY, pos=(0,0), size=(self.frame_w, CTR_HEIGHT))
        self.ctrl_panel.SetBackgroundColour('white')

        self.list_panel = wx.Panel(self.mainPanel,wx.ID_ANY, pos=(0, CTR_HEIGHT),
                                   size=(self.frame_w, self.frame_h-CTR_HEIGHT - LAST_HEIGHT))
        # self.list_panel.SetBackgroundColour('green')

        self.last_panel = wx.Panel(self.mainPanel,wx.ID_ANY, pos=(0, self.frame_h - LAST_HEIGHT),
                                   size=(self.frame_w, LAST_HEIGHT))
        self.v_layout1 = wx.BoxSizer(wx.HORIZONTAL)

        # addボタン
        self.add_button = wx.Button(self.ctrl_panel, wx.ID_ANY, '行追加', size = (BUTTON_WIDTH ,BUTTON_HEIGHT))
        self.Bind(wx.EVT_BUTTON, self.add_row_button_handler, self.add_button)  # イベントを設定
        self.v_layout1.Add(self.add_button, wx.ALL|wx.EXPAND)

        #　copyボタン
        self.copy_button = wx.Button(self.ctrl_panel, wx.ID_ANY, '行複製', size = (BUTTON_WIDTH ,BUTTON_HEIGHT))
        self.Bind(wx.EVT_BUTTON, self.copy_button_handler, self.copy_button)  # イベントを設定
        self.v_layout1.Add(self.copy_button, wx.ALL|wx.EXPAND)

        #　deleteボタン
        self.delete_button = wx.Button(self.ctrl_panel, wx.ID_ANY, '行削除', size = (BUTTON_WIDTH ,BUTTON_HEIGHT))
        self.Bind(wx.EVT_BUTTON, self.delete_button_handler, self.delete_button)  # イベントを設定
        self.v_layout1.Add(self.delete_button, wx.ALL|wx.EXPAND)

        # undoボタン
        self.undo_button = wx.Button(self.ctrl_panel, wx.ID_ANY, '元に戻す', size = (BUTTON_WIDTH ,BUTTON_HEIGHT))
        self.Bind(wx.EVT_BUTTON, self.undo_button_handler, self.undo_button)  # イベントを設定
        self.v_layout1.Add(self.undo_button, wx.ALL|wx.EXPAND)

        self.v_layout1.Add(BUTTON_WIDTH/2,BUTTON_HEIGHT)

        # Okボタン
        self.save_button = wx.Button(self.ctrl_panel, wx.ID_ANY, 'OK', size = (BUTTON_WIDTH , BUTTON_HEIGHT))
        self.Bind(wx.EVT_BUTTON, self.save_button_handler, self.save_button)  # イベントを設定
        self.v_layout1.Add(self.save_button, wx.ALL | wx.EXPAND)

        # cancelボタン
        self.cancel_button = wx.Button(self.ctrl_panel, wx.ID_ANY, 'CANCEL', size = (BUTTON_WIDTH ,BUTTON_HEIGHT))
        self.Bind(wx.EVT_BUTTON, self.cancel_button_handler, self.cancel_button)  # イベントを設定
        self.v_layout1.Add(self.cancel_button, wx.ALL|wx.EXPAND)

        self.v_layout1.Add(BUTTON_WIDTH / 2, BUTTON_HEIGHT)

        self.ctrl_panel.SetSizer(self.v_layout1)
        self.ctrl_panel.Layout()

        self.user_dic = "dic/正誤表.xlsx"
        book = openpyxl.load_workbook(self.user_dic)

        active_sheet = book.active
        self.dic = []
        for row in active_sheet.rows:
            col = []
            for cell in row:
                col.append(cell.value)
            self.dic.append(col)

        row_n = len(self.dic)
        col_n =len(self.dic[0])

        self.grid = wx.grid.Grid(self.list_panel)
        self.grid.Bind(wx.grid.EVT_GRID_CELL_CHANGED, self.cell_changed)
        self.grid.CreateGrid(row_n, col_n)

        # Set column labels.
        self.grid.SetColLabelValue(0, "誤")
        self.grid.SetColSize(0, 155)
        self.grid.SetColLabelValue(1, "正")
        self.grid.SetColSize(1, 155)

        n = 0
        for row in self.dic:
            m = 0
            for item in row:
                self.grid.SetCellValue(n, m, item)
                m += 1
            n += 1

        # Alignment.
        # self.grid.AutoSize()

        layout = wx.BoxSizer(wx.VERTICAL)
        layout.Add(self.grid, flag=wx.EXPAND | wx.ALL, border=10, proportion=1)
        self.list_panel.SetSizer(layout)
        self.list_panel.Layout()

        self.undo_n = 0
        self.undo = []
        self.add_undo()
        self.undo_button.Enabled = False
        self.save_button.Enabled = False

        # self.grid.Bind(wx.grid.EVT_GRID_CELL_RIGHT_CLICK, self.OnRightUp)
        self.grid.Bind(wx.grid.EVT_GRID_LABEL_RIGHT_CLICK, self.OnRightUp)

    def OnRightUp(self, event):

        n = self.grid.GetSelectedRows()
        if len(n) > 0:
            if not hasattr(self, "popupID1"):  # 起動後、一度だけ定義する。
                self.popupID1 = wx.NewId()
                self.popupID2 = wx.NewId()

                self.Bind(wx.EVT_MENU, self.OnPopupOne, id=self.popupID1)
                self.Bind(wx.EVT_MENU, self.OnPopupTwo, id=self.popupID2)

            menu = wx.Menu()
            menu.Append(self.popupID1, "行複製")
            menu.Append(self.popupID2, "行削除")

            self.grid.PopupMenu(menu)
            menu.Destroy()

    def OnPopupOne(self, event):
        self.row_copy()

    def OnPopupTwo(self, event):
        self.row_delete()

    def cell_changed(self, event):
        self.add_undo()

    def add_undo(self):
        self.undo_n += 1
        row_n = self.grid.GetNumberRows()
        col_n = self.grid.GetNumberCols()
        data = []
        for i in range(row_n):
            c = []
            for j in range(col_n):
                c.append(self.grid.GetCellValue(i, j))
            data.append(c)
        self.undo.append(data)
        self.undo_button.Enabled = True
        self.save_button.Enabled = True

    def save_button_handler(self, event):
        wb = openpyxl.Workbook()
        sheet = wb.worksheets[0]
        row_n = self.grid.GetNumberRows()
        col_n = self.grid.GetNumberCols()
        for i in range(row_n):
            for j in range(col_n):
                sheet.cell(column=j+1, row=i+1, value=self.grid.GetCellValue(i, j))
        wb.save(self.user_dic)
        self.Destroy()

    def cancel_button_handler(self, event):
        self.Destroy()


    def add_row_button_handler(self, event):
        n = self.grid.GetSelectedRows()
        if len(n) > 0:
            row_i = n[0]
            self.grid.InsertRows(row_i + 1)
        else:
            row_n = self.grid.GetNumberRows()
            self.grid.InsertRows(row_n)
        self.add_undo()

    def undo_button_handler(self, event):
        if self.undo_n > 1:
            row_n = self.grid.GetNumberRows()
            # for i in range(row_n):
            self.grid.ClearGrid()
            self.grid.DeleteRows(0, row_n)

            self.undo_n -= 1
            data = self.undo[self.undo_n - 1]
            self.undo.pop()

            row_n = len(data)
            col_n = len(data[0])
            for i in range(row_n):
                self.grid.InsertRows(i)
                for j in range(col_n):
                    d = data[i][j]
                    self.grid.SetCellValue(i, j, d)
            if self.undo_n > 1:
                self.undo_button.Enabled = True
                self.save_button.Enabled = True
            else:
                self.undo_button.Enabled = False
                self.save_button.Enabled = False

    def copy_button_handler(self, event):
        self.row_copy()

    def row_copy(self):
        n = self.grid.GetSelectedRows()
        if len(n) > 0:
            row_i = n[0]
            col_n = self.grid.GetNumberCols()
            d = []
            for i in range(col_n):
                d.append(self.grid.GetCellValue((row_i, i)))
            self.grid.InsertRows(row_i + 1)
            for i in range(col_n):
                self.grid.SetCellValue(row_i+1, i, d[i])
            self.add_undo()

    def delete_button_handler(self, event):
        self.row_delete()

    def row_delete(self):
        n = self.grid.GetSelectedRows()
        if len(n) > 0 :
            self.grid.DeleteRows(pos = n[0])
            self.add_undo()




if __name__ == '__main__':
    app = wx.App()
    frame1 = Errata(None, -1)
    frame1.Show()
    app.MainLoop()